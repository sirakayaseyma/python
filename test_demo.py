from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
import openpyxl

class Test_DemoClass:
    #her testten önce çağırılır.
    def setup_method(self): #genel kruluna ihtiyaç duyulabilmesine yardımcı olur.
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get("https://www.saucedemo.com/")
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)
        #günün tarihini al bu tarih ile klasör var mı kontrol et yoksa oluştur.
    #her testten sonra çağırılır
    def teardown_metod(self): #çıkış gibi biirme işlemlerinde yardımcı olr.
        self.driver.quit() #ilgili yapıyı kapatmama yardımcı olur.

    def readData(self):
        print("x")
        #setup -> test_demoFunc -> teardown
    def test_demoFunc(self):
        print("demo test")
        text = "Hello"
        assert text == "Hello"
    def test_demo2(self):
        assert True

    def getData():
        #veriyi al
        excelFile = openpyxl.load_workbook("data/invalid_login.xlsx")
        selectedSheet = excelFile["Sayfa1"]

        totalRows = selectedSheet.max_row
        data = []
        for i in range(2, totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)
        return data

        #return [("1", "1"), ("kullaniciadim" , "şifrem"), ("kodlamaio"  , "123")]
    
    # @pytest.mark.skip() # çalıştırma bu kısmı atla demek istiyor bundan sonra gelen kısmı
    @pytest.mark.parametrize("username,password" , getData()) #liste içindekit üm verilerimi çalıştırdı.
    def test_invalid_login(self,username,password):
        self.waitForElementVisible((By.ID, "user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID, "password"))
        passwordInput = self.driver.find_element(By.ID, "password")
        #en fazla 5 saniye olacak şekilde user-name id'li elementin görünmesini bekle
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"
    def waitForElementVisible(self,locator,timeout=5):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))

 